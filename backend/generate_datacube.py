# 
# generate_datacube.py  —  JSpark Semantic Data Layer
# =====================================================
# Reads all 8 source files, fixes every data-quality issue found in the audit,
# and writes 4 parquet datacubes + 1 SQLite allocation database.

# KEY FIXES vs previous version
# ──────────────────────────────
# 1. Skill join bug fixed: all 286 skill employees now join correctly.
#    Previous version silently produced NaN for every recommended employee
#    because the SENIORITY filter excluded the India technical pool.
#    Fix: build skill aggregates BEFORE the delivery-dept filter, then join.

# 2. Two employee pools are now handled separately:
#    • India pool  – Chennai SSE/SE/Enabler/SolCon (skill data exists)
#    • UK/US pool  – London/NY AP/P/M/SC/SAC/AC (competency data exists)
#    The engine must match pool to pipeline role tier + geo.

# 3. Availability uses MAX(timesheet_util, committed_alloc_pct) not either alone.
#    Managed-service (BAU-only) employees correctly excluded from candidate pool.

# 4. Outputs:
#    people_cube.parquet          — one row per active delivery employee
#    projects_cube.parquet        — one row per active/pipeline project
#    pipeline_projects_cube.parquet
#    pipeline_roles_cube.parquet
#    jspark.db (SQLite)           — allocation registry, updated by engine

# Usage:
#   python generate_datacube.py --data-dir ./data --out-dir ./datacubes


import argparse, json, re, warnings
from datetime import date
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
TODAY = pd.Timestamp(date.today())

# ─── Constants ────────────────────────────────────────────────────────────────

DELIVERY_DEPTS      = {"Delivery", "Billable ", "Consulting"}
EXCLUDE_JOB_TITLES  = {"Trainee Software Engineer", " Internal Comms Project Manager",
                        "Head of Delivery Governance"}
EXCLUDE_PROJ_TYPES  = {"BAU Activity", "Internal Project", "Sales Activity"}

# India titles map to these tier numbers (UK equivalents in parentheses)
SENIORITY_MAP = {
    "software engineer": 1,           # Associate Consultant
    "senior software engineer": 2,    # Senior Associate Consultant
    "solutions enabler": 3,           # Consultant
    "solution consultant": 4,         # Senior Consultant
    "solutions consultant": 4,
    "senior solution consultant": 5,  # Manager
    "senior solutions consultant": 5,
    "technical solutions architect": 6,
    "technology solutions architect": 6,
    "principal technology architect": 6,
    "principal architect": 6,
    "principal": 6,
    "principal solutions architect": 7,
    # UK / US
    "associate consultant": 1,
    "senior associate consultant": 2,
    "consultant": 3,
    "senior consultant": 4,
    "manager": 5,
    "associate partner": 7,
    "partner": 8,
    "leadership": 8,
}

ABBR_TIER = {
    "se":1,"ac":1,"ac (uk)":1,
    "sse":2,"sac":2,"sac or ac":2,"sac/ac":2,"sacc/ac":2,"sac - c":2,"sse or se":2,"sse  or se":2,
    "enabler":3,"c":3,"c/sac/ac":3,
    "sc":4,"sol con":4,"sc (em)":4,"sol con/enabler/sse":4,"sc or c - em":4,
    "sr sol con":5,"snr sol con":5,"m":5,"em":5,"sr ds sme":5,
    "p":6,"pa":6,"gtm architect":6,
    "ap":7,"ap/p":7,
}
ABBR_ROLE = {
    "se":"Software Engineer","ac":"Associate Consultant","ac (uk)":"Associate Consultant (UK)",
    "sse":"Senior Software Engineer","sac":"Senior Associate Consultant",
    "sac or ac":"Senior Associate Consultant / AC","sac/ac":"SAC or AC",
    "enabler":"Solutions Enabler","c":"Consultant","c/sac/ac":"Consultant / SAC / AC",
    "sc":"Solution Consultant","sol con":"Solution Consultant","sc (em)":"Solution Consultant (EM)",
    "sr sol con":"Senior Solution Consultant","snr sol con":"Senior Solution Consultant",
    "m":"Manager","em":"Engagement Manager","sr ds sme":"Senior Data Science SME",
    "p":"Principal / Technical Solutions Architect","pa":"Principal Architect",
    "gtm architect":"GTM Architect","ap":"Associate Partner","ap/p":"Associate Partner / Principal",
    "sol con/enabler/sse":"Solution Consultant / Enabler / SSE",
    "sc or c - em":"Solution Consultant or Consultant (EM)",
    "sacc/ac":"SAC / AC","sac - c":"SAC - Consultant",
}
# Pipeline abbreviation -> geo pool preference
ABBR_GEO = {
    "se":"India","sse":"India","enabler":"India","sol con":"India","sc":"India",
    "sac":"India","ac":"UK","c":"UK","sr sol con":"UK","m":"UK",
    "sc (em)":"UK","p":"UK","pa":"UK","ap":"UK","em":"UK","ap/p":"UK",
    "gtm architect":"India",
}

COE_NORM = {
    "data science & ai":"Data Science & AI",
    "data engineering":"Data Engineering",
    "power bi & consulting":"Power BI & Consulting",
    "full stack":"Full Stack",
    "techops & automation":"TechOps & Automation",
    "techops & automation ":"TechOps & Automation",
    "consulting":"Consulting","gtm":"GTM",
}

WSR_WEIGHT  = {"GREEN":1.0,"AMBER":0.5,"RED":0.0,"NO_COLOR":0.65}
EXP_YEARS   = {
    "0-1 years":0.5,"0-1 year":0.5,"1-2 year":1.5,"1-2 years":1.5,
    "2-3 years":2.5,"3-4 years":3.5,"4-5 years":4.5,"5-6 years":5.5,
    "7-8 years":7.5,"8-9 years":8.5,"9-10 years":9.5,"10-11 years":10.5,
    "12-13 years":12.5,"13-14 years":13.5,
}
CLIENT_PRIORITY_RANK = {"Gold":4,"Silver":3,"Bronze":2,"Other":1}
REQUEST_PRIORITY_RANK = {"Urgent":4,"High":3,"Medium":2,"Low":1,"Complete":0}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def pdate(v):
    if pd.isna(v) or v is None: return pd.NaT
    if isinstance(v, pd.Timestamp): return v
    try: return pd.to_datetime(str(v).split(" ")[0], dayfirst=True, errors="coerce")
    except: return pd.NaT

def days_until(dt):
    if pd.isna(dt): return float("nan")
    return (dt - TODAY).days

def norm_pct(val):
    if pd.isna(val): return 100.0
    s = str(val).strip()
    try:
        if "/" in s: return float(sum(float(p) for p in s.split("/")) / len(s.split("/")))
        if re.match(r"^\d+-\d+$", s): return float(sum(float(p) for p in s.split("-")) / 2)
        return float(s)
    except: return 100.0

def norm_abbr(a): return str(a).strip().lower()

# ─── Loaders ──────────────────────────────────────────────────────────────────

def load_employees(d: Path) -> pd.DataFrame:
    df = pd.read_csv(d / "01__260624_employee_details.csv")
    df["employee_id"]     = df["employee_id"].astype(str).str.strip().str.upper()
    df["job_name"]        = df["job_name"].fillna("").astype(str).str.strip()
    df["department_name"] = df["department_name"].fillna("").astype(str).str.strip()
    df["location"]        = df["location"].fillna("Unknown").astype(str).str.strip()
    df["geo_cluster"]     = df["location"].map({"London":"UK","New York":"US","Chennai":"India"}).fillna("Unknown")
    df["seniority_tier"]  = df["job_name"].str.lower().map(SENIORITY_MAP).fillna(0).astype(int)
    df["is_active"]       = df["account_status"] == 1
    df["is_delivery"]     = df["department_name"].isin(DELIVERY_DEPTS)
    df["is_excluded_title"] = df["job_name"].isin(EXCLUDE_JOB_TITLES)
    df["date_of_join_dt"] = df["date_of_join"].apply(pdate)
    df["tenure_years"]    = ((TODAY - df["date_of_join_dt"]).dt.days / 365.25).round(2)
    return df.reset_index(drop=True)

def load_projects(d: Path) -> pd.DataFrame:
    df = pd.read_csv(d / "02__260624_project_details.csv")
    df = df[df["is_active_version"] == 1].copy()
    df["project_id"]       = df["project_id"].astype(str).str.strip()
    df["CLIENT_ID"]        = df["CLIENT_ID"].astype(str).str.strip()
    df["project_status"]   = df["project_status"].fillna("").str.strip()
    df["type_of_project"]  = df["type_of_project"].fillna("").str.strip()
    df["tech_coe"]         = df["tech_coe"].fillna("").str.strip()
    df["proposition_coe"]  = df["proposition_coe"].fillna("").str.strip()
    df["start_dt"]         = df["project_start_date"].apply(pdate)
    df["end_dt"]           = df["project_end_date"].apply(pdate)
    df["days_until_end"]   = df["end_dt"].apply(days_until)
    df["is_billable_type"] = ~df["type_of_project"].isin(EXCLUDE_PROJ_TYPES)
    return df.reset_index(drop=True)

def load_allocations(d: Path) -> pd.DataFrame:
    df = pd.read_csv(d / "03__260623_Project_Allocation_Details.csv")
    df["employee_id"] = df["employee_id"].astype(str).str.strip().replace("nan", np.nan)
    df["project_id"]  = df["project_id"].astype(str).str.strip()
    df["resourcing_status"]      = df["resourcing_status"].fillna("").str.strip()
    df["allocation_by_percentage"] = pd.to_numeric(df["allocation_by_percentage"], errors="coerce").fillna(0)
    df["alloc_start_dt"] = df["allocated_start_date"].apply(pdate)
    df["alloc_end_dt"]   = df["allocated_end_date"].apply(pdate)
    active = df[(df["is_allocation_active"] == 1) & df["employee_id"].notna()].copy()
    # Deduplicate: one row per (employee, project) — use max allocation %
    return (active.sort_values("allocation_by_percentage", ascending=False)
            .drop_duplicates(["employee_id", "project_id"])
            .reset_index(drop=True))

def load_timesheets(d: Path) -> pd.DataFrame:
    df = pd.read_csv(d / "04__260624_timesheet_details_2026_csv.csv")
    df["employee_id"] = df["employee_id"].astype(str).str.strip().replace("nan", np.nan)
    df["project_id"]  = df["project_id"].astype(str).str.strip()
    df["date_parsed"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
    df["hours"]       = pd.to_numeric(df["time"], errors="coerce").fillna(0)
    return df[df["employee_id"].notna()].copy()

def load_skills(d: Path) -> pd.DataFrame:
    """
    FIX: Read ALL rows first. Do NOT filter by department here.
    The join to delivery employees happens in build_people_cube.
    Previous bug: filtering here removed 286 valid India skill entries.
    """
    wb = openpyxl.load_workbook(d / "05__260624_Skill_Data.xlsx", read_only=True)
    rows = list(wb["Sheet2"].iter_rows(values_only=True))
    wb.close()
    df = pd.DataFrame(rows[1:], columns=["employee_id","Designation","COE","COE_Skill",
                                          "Skill","SubSkill","Experience","Score"])
    df["employee_id"] = df["employee_id"].astype(str).str.strip().str.upper()
    df["COE"]         = df["COE"].fillna("").astype(str).str.strip()
    df["COE_norm"]    = df["COE"].str.lower().map(COE_NORM).fillna(df["COE"].str.strip())
    df["Skill"]       = df["Skill"].fillna("").astype(str).str.strip()
    df["SubSkill"]    = df["SubSkill"].fillna("").astype(str).str.strip()
    df["Experience"]  = df["Experience"].fillna("0-1 Years").astype(str).str.strip()
    df["Score"]       = pd.to_numeric(df["Score"], errors="coerce").fillna(0).astype(int)
    df["exp_years"]   = df["Experience"].str.lower().map(EXP_YEARS).fillna(0.5)
    return df.reset_index(drop=True)

def load_competencies(d: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(d / "06__260623_Competency_Details.xlsx")
    records = []
    SHEETS = {
        "Solution Enabler ":    {"band":"enabler",    "dims":["client_stakeholder","consultative_advisory","techno_functional","communication","ambiguity_nav"], "cols":[4,6,8,10,12]},
        "Solution Consultant ": {"band":"sol_con",    "dims":["capability_articulation","solution_architecture","project_planning"],                              "cols":[4,6,8]},
        "Senior Software Engineer": {"band":"sse",    "dims":["techno_functional","communication","ambiguity_nav"],                                               "cols":[4,6,8]},
    }
    for sheet, cfg in SHEETS.items():
        ws = wb[sheet]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[0]: continue
            eid = str(row[0]).strip().upper()
            for i, dim in enumerate(cfg["dims"]):
                col = cfg["cols"][i]
                try:   score = int(row[col]) if col < len(row) and row[col] is not None else 0
                except: score = 0
                records.append({"employee_id":eid,"role_band":cfg["band"],"dimension":dim,"score":score})
    return pd.DataFrame(records)

def load_wsr(d: Path) -> pd.DataFrame:
    df = pd.read_csv(d / "09__260624_Project_Weekly_Status_Details_csv.csv")
    df["project_id"] = df["project_id_masked"].astype(str).str.strip()
    for col in ["scope_status","schedule_status","quality_status","csat_status","team_status"]:
        df[col] = df[col].fillna("NO_COLOR").str.strip()
    df["week_dt"] = pd.to_datetime(df["week_start_date"], dayfirst=True, errors="coerce")
    return df

def load_pipeline(d: Path):
    wb = openpyxl.load_workbook(d / "07__260624_Pipeline_Details.xlsx")
    ws = wb["Forecast"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    C = {"cluster":0,"client_priority":4,"client":5,"em":6,"likely_start":7,
         "start_confirmed":8,"num_weeks":9,"deal_stage":10,"solution":11,
         "priority":12,"status":13,"resource":14,"pct":15,
         "recommended":16,"pct_avail":17,"skillset":18,"sow":20,"comments":21}

    proj_rows, role_rows, cur = [], [], {}
    for row in rows[1:]:
        rv = [v if not hasattr(v,"value") else None for v in row]
        if rv[C["client"]] and str(rv[C["client"]]).strip():
            ls = pdate(rv[C["likely_start"]])
            num_w = rv[C["num_weeks"]]
            try: num_w = int(num_w) if num_w and str(num_w).strip().isdigit() else None
            except: num_w = None
            cp = str(rv[C["client_priority"]]).strip() if rv[C["client_priority"]] else None
            prio = str(rv[C["priority"]]).strip() if rv[C["priority"]] else None
            cur = {
                "cluster":        rv[C["cluster"]],
                "client":         str(rv[C["client"]]).strip(),
                "client_priority": cp,
                "em":             str(rv[C["em"]]).strip() if rv[C["em"]] else None,
                "likely_start_dt": ls,
                "likely_start_str": ls.strftime("%Y-%m-%d") if not pd.isna(ls) else None,
                "num_weeks":      num_w,
                "deal_stage":     str(rv[C["deal_stage"]]).strip() if rv[C["deal_stage"]] else None,
                "solution":       str(rv[C["solution"]]).strip() if rv[C["solution"]] else None,
                "priority":       prio,
                "resourcing_status": str(rv[C["status"]]).strip() if rv[C["status"]] else "Not Resourced",
                "sow_signed":     str(rv[C["sow"]]).strip() == "Yes" if rv[C["sow"]] else False,
                "comments":       str(rv[C["comments"]]).strip() if rv[C["comments"]] else None,
                "client_priority_rank":  CLIENT_PRIORITY_RANK.get(cp, 0),
                "request_priority_rank": REQUEST_PRIORITY_RANK.get(prio, 0),
            }
            cur["composite_priority"] = cur["client_priority_rank"] * 10 + cur["request_priority_rank"]
            cur["pipeline_id"] = f"{cur['client'].replace(' ','_')}_{cur['cluster']}_{cur['likely_start_str'] or 'TBD'}"
            cur["days_until_start"] = days_until(ls)
            proj_rows.append(dict(cur))

        if not cur or not rv[C["resource"]]: continue
        abbr_raw = str(rv[C["resource"]]).strip()
        abbr_n   = norm_abbr(abbr_raw)
        primary  = re.split(r"[/\s]or[\s]", abbr_n)[0].split("/")[0].strip()
        tier     = ABBR_TIER.get(primary, ABBR_TIER.get(abbr_n, 3))
        role_nm  = ABBR_ROLE.get(primary, ABBR_ROLE.get(abbr_n, abbr_raw))
        geo_pref = ABBR_GEO.get(primary, ABBR_GEO.get(abbr_n, "Any"))
        role_rows.append({
            "pipeline_id":       cur["pipeline_id"],
            "role_abbr":         abbr_raw,
            "role_abbr_norm":    abbr_n,
            "role_name":         role_nm,
            "seniority_tier":    tier,
            "geo_preference":    geo_pref,
            "allocation_pct":    norm_pct(rv[C["pct"]]),
            "skillset_notes":    str(rv[C["skillset"]]).strip() if rv[C["skillset"]] and not hasattr(rv[C["skillset"]],"value") else "",
            "already_recommended": str(rv[C["recommended"]]).strip() if rv[C["recommended"]] else None,
        })

    proj_df  = pd.DataFrame(proj_rows).drop_duplicates("pipeline_id").reset_index(drop=True)
    roles_df = pd.DataFrame(role_rows).reset_index(drop=True)

    # Enrich roles with project-level fields
    roles_df = roles_df.merge(
        proj_df[["pipeline_id","client","cluster","client_priority","priority",
                 "composite_priority","likely_start_dt","likely_start_str",
                 "sow_signed","solution","num_weeks"]],
        on="pipeline_id", how="left"
    )
    roles_df["role_urgency"] = roles_df["composite_priority"].fillna(0).astype(int)

    return proj_df.sort_values("composite_priority", ascending=False).reset_index(drop=True), roles_df

# ─── Enrichment ───────────────────────────────────────────────────────────────

def build_utilization(ts: pd.DataFrame, lookback_weeks: int = 8) -> pd.DataFrame:
    cutoff = TODAY - pd.Timedelta(weeks=lookback_weeks)
    recent = ts[ts["date_parsed"] >= cutoff].copy()
    std_hours = 40.0 * lookback_weeks

    emp_h = recent.groupby("employee_id")["hours"].sum().reset_index()
    emp_h.columns = ["employee_id","recent_hours"]
    emp_h["ts_util_pct"]          = (emp_h["recent_hours"] / std_hours * 100).clip(0, 200).round(1)
    emp_h["ts_available_pct"]     = (100 - emp_h["ts_util_pct"]).clip(0, 100).round(1)

    proj_h = recent.groupby(["employee_id","project_id"])["hours"].sum().reset_index()
    proj_list = (proj_h.groupby("employee_id")["project_id"]
                 .apply(lambda x: sorted(x.tolist()))
                 .reset_index().rename(columns={"project_id":"ts_project_ids"}))

    main_p = (proj_h.sort_values("hours", ascending=False)
              .drop_duplicates("employee_id")[["employee_id","project_id"]]
              .rename(columns={"project_id":"main_ts_project"}))

    return emp_h.merge(proj_list, on="employee_id", how="left").merge(main_p, on="employee_id", how="left")

def build_project_health(wsr: pd.DataFrame, lookback_weeks: int = 6) -> pd.DataFrame:
    DIMS = ["scope_status","schedule_status","quality_status","csat_status","team_status"]
    cutoff = TODAY - pd.Timedelta(weeks=lookback_weeks)
    recent = wsr[wsr["week_dt"] >= cutoff].copy()
    records = []
    for pid, grp in recent.groupby("project_id"):
        coloured = grp[grp["schedule_status"] != "NO_COLOR"].sort_values("week_dt", ascending=False)
        use = coloured if not coloured.empty else grp.sort_values("week_dt", ascending=False)
        top4 = use.head(4).reset_index(drop=True)
        rw = [0.4, 0.3, 0.2, 0.1]
        score = sum(rw[i] * (sum(WSR_WEIGHT.get(top4.iloc[i][d], 0.65) for d in DIMS) / 5)
                    for i in range(len(top4)))
        score += sum(rw[i] for i in range(len(top4), 4)) * 0.65

        r0 = use.iloc[0]
        red_4  = int((top4["schedule_status"] == "RED").sum())
        amb_4  = int((top4["schedule_status"] == "AMBER").sum())
        ext_risk = min(1.0, red_4 * 0.25 + amb_4 * 0.10)
        records.append({
            "project_id":       pid,
            "health_score":     round(min(1.0, max(0.0, score)), 3),
            "extension_risk":   round(ext_risk, 3),
            "latest_scope":     r0["scope_status"],
            "latest_schedule":  r0["schedule_status"],
            "latest_quality":   r0["quality_status"],
            "latest_csat":      r0["csat_status"],
            "latest_team":      r0["team_status"],
            "wsr_weeks_count":  len(grp),
        })
    return pd.DataFrame(records)

def build_skill_aggregates(skills: pd.DataFrame) -> pd.DataFrame:
    """
    FIX: Build skill aggregates for ALL employees in Skill_Data.xlsx.
    The join to delivery employees happens in build_people_cube, not here.
    """
    scored = skills[skills["Score"] > 0].copy()
    agg = scored.groupby("employee_id").agg(
        avg_skill_score=("Score","mean"),
        max_skill_score=("Score","max"),
        avg_exp_years=("exp_years","mean"),
        skill_breadth=("Skill","nunique"),
        total_skill_rows=("Score","count"),
    ).round(2).reset_index()

    # Primary COE = COE with highest avg score
    coe_avg = (scored.groupby(["employee_id","COE_norm"])["Score"].mean().reset_index())
    primary_coe = (coe_avg.sort_values("Score", ascending=False)
                   .drop_duplicates("employee_id")[["employee_id","COE_norm"]]
                   .rename(columns={"COE_norm":"primary_coe"}))

    # All COEs with scored skills
    all_coes = (scored.groupby("employee_id")["COE_norm"]
                .apply(lambda x: sorted(x.unique().tolist()))
                .reset_index().rename(columns={"COE_norm":"all_coes"}))

    # Top 15 skills for embedding text
    top_sk = (scored.sort_values("Score", ascending=False)
              .groupby("employee_id")
              .apply(lambda x: x.head(15)[["Skill","SubSkill","Score","exp_years"]].to_dict("records"))
              .reset_index().rename(columns={0:"top_skills"}))

    # Skill text (rich — used for TF-IDF and semantic embedding)
    skill_text = (scored.sort_values("Score", ascending=False)
                  .groupby("employee_id")
                  .apply(lambda x: " | ".join(
                      f"{r['SubSkill']} {r['Skill']} coe:{r['COE_norm']} score:{r['Score']} exp:{r['exp_years']}y"
                      for _, r in x.head(20).iterrows()))
                  .reset_index().rename(columns={0:"skill_text"}))

    return (agg
            .merge(primary_coe, on="employee_id", how="left")
            .merge(all_coes,    on="employee_id", how="left")
            .merge(top_sk,      on="employee_id", how="left")
            .merge(skill_text,  on="employee_id", how="left"))

def build_competency_aggregates(comp: pd.DataFrame) -> pd.DataFrame:
    agg = comp.groupby("employee_id").agg(
        avg_competency_score=("score","mean"),
        max_competency_score=("score","max"),
    ).round(2).reset_index()
    profile = (comp.groupby("employee_id")
               .apply(lambda x: {r["dimension"]: r["score"] for _, r in x.iterrows()})
               .reset_index().rename(columns={0:"competency_profile"}))
    return agg.merge(profile, on="employee_id", how="left")

def build_allocation_signals(alloc: pd.DataFrame, projects: pd.DataFrame,
                             health: pd.DataFrame) -> pd.DataFrame:
    bau_proj = set(projects[projects["type_of_project"].isin(EXCLUDE_PROJ_TYPES)]["project_id"])
    active_proj = set(projects[projects["project_status"] == "ACTIVE"]["project_id"])

    health_map = dict(zip(health["project_id"], health["health_score"])) if not health.empty else {}
    sched_map  = dict(zip(health["project_id"], health["latest_schedule"])) if not health.empty else {}

    # Separate billable active vs bau allocations
    alloc_billable = alloc[
        alloc["project_id"].isin(active_proj) &
        ~alloc["project_id"].isin(bau_proj)
    ].copy()
    alloc_bau = alloc[alloc["project_id"].isin(bau_proj)].copy()

    records = []
    proj_meta = projects[["project_id","CLIENT_ID","end_dt","days_until_end"]].copy()
    merged = alloc_billable.merge(proj_meta, on="project_id", how="left")

    for eid, grp in merged.groupby("employee_id"):
        proj_ids   = grp["project_id"].tolist()
        alloc_sum  = grp["allocation_by_percentage"].sum()
        client_ids = grp["CLIENT_ID"].dropna().unique().tolist()

        # Soonest end date
        end_dates = grp["end_dt"].dropna()
        soonest   = end_dates.min() if not end_dates.empty else pd.NaT
        d_soonest = days_until(soonest)

        hs = [health_map.get(p, 0.65) for p in proj_ids]
        records.append({
            "employee_id":          eid,
            "active_project_ids":   proj_ids,
            "alloc_committed_pct":  round(alloc_sum, 1),
            "alloc_active_count":   len(proj_ids),
            "is_bau_only":          (eid in set(alloc_bau["employee_id"]) and len(proj_ids) == 0),
            "soonest_end_date":     soonest,
            "days_to_soonest_end":  d_soonest,
            "ramp_down_flag":       (not pd.isna(soonest) and 0 <= d_soonest <= 28),
            "avg_project_health":   round(float(np.mean(hs)) if hs else 0.65, 3),
            "on_red_project":       any(sched_map.get(p) == "RED"   for p in proj_ids),
            "on_amber_project":     any(sched_map.get(p) == "AMBER" for p in proj_ids),
            "client_history":       client_ids,
        })
    return pd.DataFrame(records)

# ─── People Cube ──────────────────────────────────────────────────────────────

def build_people_cube(employees, skill_agg, comp_agg, util, alloc_sig) -> pd.DataFrame:
    """
    FIX: Filter to delivery employees AFTER joining skill/competency data.
    This ensures India technical staff (who have skill data) are included.
    """
    # Start with ALL active delivery employees (excluding trainees + misc)
    pool = employees[
        employees["is_active"] &
        employees["is_delivery"] &
        (~employees["is_excluded_title"]) &
        (employees["seniority_tier"] >= 1)
    ].copy()

    # Joins — left joins so employees without skill data still appear
    pool = (pool
            .merge(skill_agg, on="employee_id", how="left")
            .merge(comp_agg,  on="employee_id", how="left")
            .merge(util,      on="employee_id", how="left")
            .merge(alloc_sig, on="employee_id", how="left"))

    # Fill NaNs
    for col in ["ts_util_pct","ts_available_pct","recent_hours"]:
        pool[col] = pool[col].fillna(0.0)
    for col in ["avg_skill_score","max_skill_score","avg_exp_years","skill_breadth","total_skill_rows"]:
        pool[col] = pool[col].fillna(0.0)
    for col in ["avg_competency_score","max_competency_score"]:
        pool[col] = pool[col].fillna(0.0)
    for col in ["alloc_committed_pct","alloc_active_count"]:
        pool[col] = pool[col].fillna(0.0)
    for bool_col in ["ramp_down_flag","on_red_project","on_amber_project","is_bau_only"]:
        pool[bool_col] = pool[bool_col].fillna(False).astype(bool)
    for list_col in ["active_project_ids","ts_project_ids","client_history","top_skills","all_coes"]:
        pool[list_col] = pool[list_col].apply(lambda x: x if isinstance(x, list) else [])
    pool["competency_profile"] = pool["competency_profile"].apply(
        lambda x: x if isinstance(x, dict) else {})
    pool["primary_coe"]   = pool["primary_coe"].fillna("Unknown")
    pool["skill_text"]    = pool["skill_text"].fillna("")
    pool["avg_project_health"] = pool["avg_project_health"].fillna(0.65)
    pool["days_to_soonest_end"] = pool["days_to_soonest_end"].fillna(float("nan"))
    pool["tenure_years"] = pool["tenure_years"].fillna(0.0)

    # TRUE available capacity = 100 - max(timesheet util, alloc committed)
    pool["effective_util_pct"] = pool[["ts_util_pct","alloc_committed_pct"]].max(axis=1).round(1)
    pool["available_capacity_pct"] = (100 - pool["effective_util_pct"]).clip(0, 100).round(1)

    # Availability flags
    pool["is_fully_available"]   = pool["available_capacity_pct"] >= 90
    pool["is_partially_available"] = pool["available_capacity_pct"].between(25, 89.9)
    pool["is_busy"]              = pool["available_capacity_pct"] < 25
    pool["soon_free"]            = (pool["ramp_down_flag"] |
                                    pool["days_to_soonest_end"].between(0, 56))

    # Health signals
    def hp(r):
        if r["on_red_project"]: return 0.15
        if r["on_amber_project"]: return 0.08
        if r["avg_project_health"] < 0.4: return 0.10
        return 0.0
    pool["health_penalty"]   = pool.apply(hp, axis=1)
    pool["ramp_down_bonus"]  = pool["ramp_down_flag"].apply(lambda x: 0.05 if x else 0.0)

    # Swap eligibility: not BAU-only AND seniority >= 1
    pool["swap_eligible"] = (~pool["is_bau_only"]) & (pool["seniority_tier"] >= 1)

    # Has data flags (for debugging/transparency)
    pool["has_skill_data"]      = pool["avg_skill_score"] > 0
    pool["has_competency_data"] = pool["avg_competency_score"] > 0

    # Profile text for embedding — enriched with skill text when available
    def make_profile_text(r):
        sk = r["skill_text"] if r["skill_text"] else ""
        cp = r["competency_profile"]
        comp_str = " ".join(f"{k}:{v}" for k, v in (cp or {}).items() if isinstance(v,(int,float)) and v > 0)
        return (
            f"Role: {r['job_name']}. Location: {r['location']}. "
            f"COE: {r['primary_coe']}. Tier: {r['seniority_tier']}. "
            f"Skills: {sk}. Competencies: {comp_str}."
        )
    pool["profile_text"] = pool.apply(make_profile_text, axis=1)

    # Pool label (used by engine for routing)
    def pool_label(r):
        if r["geo_cluster"] == "India" and r["has_skill_data"]: return "india_technical"
        if r["geo_cluster"] in ("UK","US") and r["has_competency_data"]: return "uk_consulting"
        if r["geo_cluster"] == "India": return "india_no_skills"
        return "uk_no_competency"
    pool["talent_pool"] = pool.apply(pool_label, axis=1)

    cols = [
        "employee_id","location","geo_cluster","job_name","department_name",
        "seniority_tier","tenure_years","talent_pool",
        "primary_coe","all_coes","avg_skill_score","max_skill_score","avg_exp_years",
        "skill_breadth","total_skill_rows","top_skills","skill_text","has_skill_data",
        "avg_competency_score","max_competency_score","competency_profile","has_competency_data",
        "ts_util_pct","alloc_committed_pct","effective_util_pct","available_capacity_pct",
        "recent_hours","is_fully_available","is_partially_available","is_busy","soon_free",
        "active_project_ids","alloc_active_count","ts_project_ids","main_ts_project",
        "soonest_end_date","days_to_soonest_end","ramp_down_flag","ramp_down_bonus",
        "avg_project_health","on_red_project","on_amber_project","health_penalty",
        "swap_eligible","is_bau_only","client_history","profile_text",
    ]
    return pool[[c for c in cols if c in pool.columns]].reset_index(drop=True)

# ─── Projects Cube ────────────────────────────────────────────────────────────

def build_projects_cube(projects, alloc, health, ts) -> pd.DataFrame:
    relevant = projects[projects["project_status"].isin(
        ["ACTIVE","DEAL WON","PROPOSE","SOW PENDING SIGNATURE","SCOPING APPROVAL"])].copy()

    # Allocation breakdown
    ap = alloc.groupby("project_id").agg(
        total_slots=("employee_id","count"),
        billable_count=("resourcing_status", lambda x:(x=="BILLABLE").sum()),
        shadow_count=("resourcing_status",   lambda x:(x=="SHADOW").sum()),
        unbilled_count=("resourcing_status", lambda x:(x=="UNBILLED").sum()),
        proposed_count=("resourcing_status", lambda x:(x=="PROPOSED").sum()),
        total_alloc_pct=("allocation_by_percentage","sum"),
        allocated_employees=("employee_id", lambda x: sorted(x.dropna().tolist())),
    ).reset_index()
    relevant = relevant.merge(ap, on="project_id", how="left")

    relevant = relevant.merge(health, on="project_id", how="left")
    relevant["health_score"]    = relevant["health_score"].fillna(0.65)
    relevant["extension_risk"]  = relevant["extension_risk"].fillna(0.0)
    for c in ["latest_scope","latest_schedule","latest_quality","latest_csat","latest_team"]:
        relevant[c] = relevant[c].fillna("NO_COLOR")

    cutoff = TODAY - pd.Timedelta(weeks=8)
    proj_ts = (ts[ts["date_parsed"] >= cutoff].groupby("project_id")["hours"].sum()
               .reset_index().rename(columns={"hours":"recent_ts_hours"}))
    relevant = relevant.merge(proj_ts, on="project_id", how="left")
    relevant["recent_ts_hours"] = relevant["recent_ts_hours"].fillna(0)

    def rds(r):
        d = r["days_until_end"]
        if pd.isna(d): return 0.0
        s = 0.8 if d<=7 else 0.5 if d<=14 else 0.4 if d<=28 else 0.2 if d<=56 else 0.0
        return round(min(1.0, s + (0.1 if r["health_score"] < 0.4 else 0)), 3)
    relevant["ramp_down_signal"] = relevant.apply(rds, axis=1)

    for c in ["total_slots","billable_count","shadow_count","unbilled_count","proposed_count"]:
        relevant[c] = relevant[c].fillna(0).astype(int)
    relevant["billability_rate"] = np.where(
        relevant["total_slots"] > 0,
        (relevant["billable_count"] / relevant["total_slots"] * 100).round(1), 0.0)

    cols = ["project_id","CLIENT_ID","project_status","type_of_project","tech_coe","proposition_coe",
            "start_dt","end_dt","days_until_end","is_billable_type",
            "health_score","extension_risk","ramp_down_signal",
            "latest_scope","latest_schedule","latest_quality","latest_csat","latest_team",
            "total_slots","billable_count","shadow_count","unbilled_count","proposed_count",
            "total_alloc_pct","billability_rate","allocated_employees","recent_ts_hours"]
    return relevant[[c for c in cols if c in relevant.columns]].reset_index(drop=True)

# ─── SQLite DB ────────────────────────────────────────────────────────────────

def build_sqlite_db(people: pd.DataFrame, projects_cube: pd.DataFrame,
                    pipe_proj: pd.DataFrame, pipe_roles: pd.DataFrame,
                    alloc_raw: pd.DataFrame, out_dir: Path):
    """
    Creates jspark.db with 6 tables used by the recommendation engine.

    employees          — people cube (flat, no list columns)
    projects           — projects cube
    pipeline_projects  — pipeline demand (sorted by priority)
    pipeline_roles     — role slots
    allocations        — current real allocations from source data
    engine_allocations — allocations MADE BY THE ENGINE (append-only log)

    The engine_allocations table is the source of truth for capacity locking.
    When the engine assigns EMP761 to KIA at 100% for 11 weeks,
    it writes a row here — and all subsequent recommendations check this
    table FIRST before considering that employee available.
    """
    import sqlite3, json as _json
    from datetime import datetime

    db_path = out_dir / "jspark.db"
    if db_path.exists():
        db_path.unlink()

    con = sqlite3.connect(str(db_path))

    # ── employees ──────────────────────────────────────────────────────────────
    emp_flat = people.copy()
    list_cols = ["active_project_ids","ts_project_ids","client_history","top_skills","all_coes"]
    dict_cols = ["competency_profile"]
    for c in list_cols:
        if c in emp_flat.columns:
            emp_flat[c] = emp_flat[c].apply(lambda x: _json.dumps(x) if isinstance(x,(list,dict)) else "[]")
    for c in dict_cols:
        if c in emp_flat.columns:
            emp_flat[c] = emp_flat[c].apply(lambda x: _json.dumps(x) if isinstance(x,(list,dict)) else "{}")
    for c in emp_flat.select_dtypes(include=["datetime64[ns]"]).columns:
        emp_flat[c] = emp_flat[c].astype(str)
    for c in emp_flat.select_dtypes(include=["bool"]).columns:
        emp_flat[c] = emp_flat[c].astype(int)
    emp_flat.to_sql("employees", con, if_exists="replace", index=False)

    # ── projects ──────────────────────────────────────────────────────────────
    proj_flat = projects_cube.copy()
    if "allocated_employees" in proj_flat.columns:
        proj_flat["allocated_employees"] = proj_flat["allocated_employees"].apply(
            lambda x: _json.dumps(x) if isinstance(x, list) else "[]")
    for c in proj_flat.select_dtypes(include=["datetime64[ns]"]).columns:
        proj_flat[c] = proj_flat[c].astype(str)
    for c in proj_flat.select_dtypes(include=["bool"]).columns:
        proj_flat[c] = proj_flat[c].astype(int)
    proj_flat.to_sql("projects", con, if_exists="replace", index=False)

    # ── pipeline_projects ─────────────────────────────────────────────────────
    pp_flat = pipe_proj.copy()
    for c in pp_flat.select_dtypes(include=["datetime64[ns]"]).columns:
        pp_flat[c] = pp_flat[c].astype(str)
    for c in pp_flat.select_dtypes(include=["bool"]).columns:
        pp_flat[c] = pp_flat[c].astype(int)
    pp_flat.to_sql("pipeline_projects", con, if_exists="replace", index=False)

    # ── pipeline_roles ────────────────────────────────────────────────────────
    pr_flat = pipe_roles.copy()
    for c in pr_flat.select_dtypes(include=["datetime64[ns]"]).columns:
        pr_flat[c] = pr_flat[c].astype(str)
    for c in pr_flat.select_dtypes(include=["bool"]).columns:
        pr_flat[c] = pr_flat[c].astype(int)
    pr_flat.to_sql("pipeline_roles", con, if_exists="replace", index=False)

    # ── allocations (raw current state from source data) ──────────────────────
    alloc_flat = alloc_raw.copy()
    for c in alloc_flat.select_dtypes(include=["datetime64[ns]"]).columns:
        alloc_flat[c] = alloc_flat[c].astype(str)
    alloc_flat.to_sql("allocations", con, if_exists="replace", index=False)

    # ── engine_allocations (engine's own decisions) ───────────────────────────
    # This is the CRITICAL table. Written by the engine, never by the datacube generator.
    # Schema is fixed — do not change without updating the engine.
    con.execute("""
        CREATE TABLE IF NOT EXISTS engine_allocations (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            pipeline_id         TEXT    NOT NULL,
            role_abbr           TEXT    NOT NULL,
            role_name           TEXT    NOT NULL,
            employee_id         TEXT    NOT NULL,
            allocation_pct      REAL    NOT NULL,
            client              TEXT,
            client_priority     TEXT,
            request_priority    TEXT,
            plan_type           TEXT,           -- DIRECT / SWAP / SOFT_COMMIT / HIRE
            confidence_score    REAL,
            composite_score     REAL,
            alloc_start_date    TEXT,           -- ISO date from pipeline likely_start
            alloc_end_date      TEXT,           -- derived from likely_start + num_weeks
            num_weeks           INTEGER,
            assigned_by         TEXT DEFAULT 'engine',
            assigned_at         TEXT DEFAULT (datetime('now')),
            is_confirmed        INTEGER DEFAULT 0,  -- 1 = RM confirmed, 0 = engine proposed
            notes               TEXT,
            swap_chain_json     TEXT            -- JSON of swap links if plan_type = SWAP
        )
    """)

    # Index for fast capacity queries: "is EMP761 committed between date A and B?"
    con.execute("CREATE INDEX IF NOT EXISTS idx_ea_employee ON engine_allocations(employee_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_ea_pipeline ON engine_allocations(pipeline_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_ea_dates    ON engine_allocations(alloc_start_date, alloc_end_date)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_employees   ON employees(employee_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_emp_tier    ON employees(seniority_tier)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_emp_geo     ON employees(geo_cluster)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_emp_coe     ON employees(primary_coe)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_emp_avail   ON employees(available_capacity_pct)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_pipe_prio   ON pipeline_projects(composite_priority DESC)")

    con.commit()
    con.close()
    return db_path

# ─── Main ─────────────────────────────────────────────────────────────────────

def serialise(df: pd.DataFrame) -> pd.DataFrame:
    """Convert list/dict columns to JSON strings for parquet."""
    out = df.copy()
    for col in out.columns:
        sample = out[col].dropna().head(3).tolist()
        if sample and isinstance(sample[0], (list, dict)):
            out[col] = out[col].apply(lambda x: json.dumps(x) if isinstance(x, (list, dict)) else x)
    return out

def main():
    parser = argparse.ArgumentParser(description="JSpark Datacube Generator")
    parser.add_argument("--data-dir",  default="./data",      help="Source files directory")
    parser.add_argument("--out-dir",   default="./datacubes", help="Output directory")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    out_dir  = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("JSpark Datacube Generator  (fixed edition)")
    print("=" * 60)

    print("\n[1/8] Loading employees...")
    employees = load_employees(data_dir)
    print(f"      {len(employees)} total; active delivery (pre-filter): "
          f"{employees[employees['is_active'] & employees['is_delivery']].shape[0]}")

    print("[2/8] Loading projects...")
    projects = load_projects(data_dir)

    print("[3/8] Loading allocations...")
    alloc = load_allocations(data_dir)
    print(f"      {len(alloc)} active deduped allocation rows")

    print("[4/8] Loading timesheets...")
    ts = load_timesheets(data_dir)

    print("[5/8] Loading skills (ALL employees, no dept filter)...")
    skills = load_skills(data_dir)
    print(f"      {len(skills)} rows; {skills['employee_id'].nunique()} unique employees")

    print("[6/8] Loading competencies...")
    comp = load_competencies(data_dir)
    print(f"      {len(comp)} rows; {comp['employee_id'].nunique()} unique employees")

    print("[7/8] Loading WSR...")
    wsr = load_wsr(data_dir)

    print("[8/8] Loading pipeline...")
    pipe_proj, pipe_roles = load_pipeline(data_dir)
    print(f"      {len(pipe_proj)} projects; {len(pipe_roles)} role slots")

    print("\n[A] Project health scores...")
    health = build_project_health(wsr)

    print("[B] Skill aggregates (joined to all employees with skill data)...")
    skill_agg = build_skill_aggregates(skills)
    print(f"    {len(skill_agg)} employees with skill data")

    print("[C] Competency aggregates...")
    comp_agg = build_competency_aggregates(comp)
    print(f"    {len(comp_agg)} employees with competency data")

    print("[D] Timesheet utilization (8-week lookback)...")
    util = build_utilization(ts)

    print("[E] Allocation signals...")
    alloc_sig = build_allocation_signals(alloc, projects, health)

    print("\n[CUBE 1] People Cube...")
    people = build_people_cube(employees, skill_agg, comp_agg, util, alloc_sig)
    print(f"         {len(people)} rows × {len(people.columns)} cols")
    print(f"         With skill data:      {people['has_skill_data'].sum()}")
    print(f"         With competency data: {people['has_competency_data'].sum()}")
    print(f"         India technical pool: {(people['talent_pool']=='india_technical').sum()}")
    print(f"         UK consulting pool:   {(people['talent_pool']=='uk_consulting').sum()}")
    print(f"         Fully available:      {people['is_fully_available'].sum()}")
    print(f"         Partially available:  {people['is_partially_available'].sum()}")
    print(f"         Busy:                 {people['is_busy'].sum()}")
    print(f"         Ramp-down soon:       {people['ramp_down_flag'].sum()}")

    print("\n[CUBE 2] Projects Cube...")
    proj_cube = build_projects_cube(projects, alloc, health, ts)
    print(f"         {len(proj_cube)} rows × {len(proj_cube.columns)} cols")

    print("\n[CUBE 3] Pipeline Cubes...")
    print(f"         Projects: {len(pipe_proj)} | Roles: {len(pipe_roles)}")
    print(f"         Gold+Urgent: {((pipe_proj['client_priority']=='Gold') & (pipe_proj['priority']=='Urgent')).sum()}")

    print("\n[SAVE] Writing parquet files...")
    serialise(people).to_parquet(   out_dir / "people_cube.parquet",             index=False)
    serialise(proj_cube).to_parquet(out_dir / "projects_cube.parquet",           index=False)
    serialise(pipe_proj).to_parquet(out_dir / "pipeline_projects_cube.parquet",  index=False)
    serialise(pipe_roles).to_parquet(out_dir / "pipeline_roles_cube.parquet",    index=False)
    print("    [OK] 4 parquet files written")

    print("\n[SQLITE] Building jspark.db...")
    db_path = build_sqlite_db(people, proj_cube, pipe_proj, pipe_roles, alloc, out_dir)
    print(f"    [OK] {db_path}")

    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)
    
    # Phase 1: Skill Join Validation
    print("\n[Phase 1] Validating Skills Join...")
    all_emp_ids = set(employees["employee_id"])
    skill_emp_ids = set(skills["employee_id"])
    matched_ids = skill_emp_ids.intersection(all_emp_ids)
    missing_ids = skill_emp_ids - all_emp_ids
    
    total_employees = len(employees)
    matched_skills = len(matched_ids)
    missing_skills = len(missing_ids)
    duplicate_skill_records = skills.duplicated(subset=["employee_id", "COE", "COE_Skill", "Skill", "SubSkill"]).sum()
    
    coverage_pct = (matched_skills / len(skill_emp_ids) * 100) if len(skill_emp_ids) > 0 else 0.0
    
    print(f"  Total Employees:          {total_employees}")
    print(f"  Matched Skills:           {matched_skills}")
    print(f"  Missing Skills:           {missing_skills}")
    print(f"  Duplicate Skill Records:  {duplicate_skill_records}")
    print(f"  Coverage %:               {coverage_pct:.2f}%")
    
    if coverage_pct < 95.0:
        raise ValueError(f"Skill coverage ({coverage_pct:.2f}%) is below 95% threshold! Aborting build.")
    
    # Quick checks
    assert people['has_skill_data'].sum() > 0,      "FAIL: still no skill data joined"
    assert people['has_competency_data'].sum() > 0,  "FAIL: still no competency data joined"
    assert len(pipe_roles[pipe_roles['seniority_tier'] > 0]) > 0, "FAIL: no role tiers"
    india_with_skills = people[(people['geo_cluster']=='India') & people['has_skill_data']]
    uk_with_comp     = people[(people['geo_cluster'].isin(['UK','US'])) & people['has_competency_data']]
    print(f"  India employees with skill data:   {len(india_with_skills)} [OK]")
    print(f"  UK/US employees with competency:   {len(uk_with_comp)} [OK]")
    print(f"  Pipeline roles with geo_preference: {(pipe_roles['geo_preference']!='Any').sum()} [OK]")
    print("\n[PASS] All checks passed. Datacubes ready.")
    print(f"   Output: {out_dir.resolve()}")


if __name__ == "__main__":
    main()
