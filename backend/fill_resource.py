import os
import sys
import subprocess
from pathlib import Path

# Resolve directories
backend_dir = Path(__file__).resolve().parent
venv_python = backend_dir / ".venv" / "Scripts" / "python.exe"

# Check if we are already running inside the virtual environment
is_in_venv = (sys.prefix != sys.base_prefix) or (backend_dir / ".venv" / "lib").exists() and (sys.executable.lower().startswith(str(backend_dir).lower()))

# If virtual env python exists and we are not currently running in it, re-execute
if venv_python.exists() and not is_in_venv:
    # Set encoding to prevent terminal cp1252 errors on Windows
    env = dict(os.environ)
    env["PYTHONIOENCODING"] = "utf-8"
    
    # Re-run the current script using venv python
    args = [str(venv_python)] + sys.argv
    result = subprocess.run(args, env=env)
    sys.exit(result.returncode)

# ─────────────────────────────────────────────────────────────────────────────
# Actual script logic starts here (runs only when inside virtual environment)
# ─────────────────────────────────────────────────────────────────────────────
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from copy import copy
load_dotenv()

sys.path.insert(0, str(backend_dir))

from engine.recommend import RecommendationEngine
from engine.config import DEFAULT_CONFIG

def generate_reasoning(engine, project_row, role_plan, recommended_option, req_skills) -> str:
    import os
    import httpx
    from openai import OpenAI

    client_priority = project_row.get("client_priority", "Bronze")
    priority = project_row.get("priority", "Medium")
    solution = project_row.get("solution", "")

    role_name = role_plan.role_name
    req_tier = role_plan.seniority_tier
    req_pct = role_plan.required_pct

    opt = recommended_option
    is_hire = opt.plan_type == "D_HIRE" if opt else False
    is_extend = opt.plan_type == "E_EXTEND_START" if opt else False

    has_skills = bool(req_skills and len(str(req_skills).strip()) > 2)

    if is_hire:
        decision_desc = f"HIRE REQUIRED (No internal candidates available at Tier {req_tier})"
    elif is_extend:
        decision_desc = f"EXTEND START DATE (Wait list / delay project recommended)"
    else:
        if not has_skills:
            decision_desc = (
                f"Allocate employee {opt.recommended_employee_id if opt else 'N/A'} ({opt.job_name if opt else 'N/A'}, "
                f"Tier {opt.seniority_tier if opt else 'N/A'}, located in {opt.location if opt else 'N/A'}) "
                f"for {req_pct}% of their capacity. "
                f"Justification: They are allocated simply because they are available and no skillset constraints were given."
            )
        else:
            sim = opt.score_breakdown.get("semantic_similarity", 0.0) if opt else 0.0
            decision_desc = (
                f"Allocate employee {opt.recommended_employee_id if opt else 'N/A'} ({opt.job_name if opt else 'N/A'}, "
                f"Tier {opt.seniority_tier if opt else 'N/A'}, located in {opt.location if opt else 'N/A'}) "
                f"for {req_pct}% of their capacity. "
                f"Justification: They have a semantic match similarity score of {sim:.2f}."
            )

    prompt = f"""
System Decision for Project Staffing Slot:
- Project ID: {project_row.get('pipeline_id')}
- Client: {project_row.get('client')} (Priority: {client_priority})
- Solution Area: {solution}
- Role Requested: {role_name} (Required Seniority Tier: {req_tier}, Allocation: {req_pct}%)
- System Decision: {decision_desc}

Task: Write a concise, 1-2 sentence justification note explaining the decision.
{f"Explain that they are allocated simply because they are available and no skillset constraints were given." if not has_skills else f"Describe how good their fit is and highlight their semantic similarity score."}
Rule: Do not change the decision or propose other candidates. Your output must only contain the justification text itself. No prefixes, no markup, no quotes.
"""

    # First attempt: OpenRouter
    openrouter_api_key = os.getenv("OPENROUTER_API_KEY", "")
    if openrouter_api_key and getattr(engine, "use_openrouter", False):
        try:
            print("        Trying OpenRouter...")
            client = OpenAI(
                base_url="https://openrouter.ai/api/v1",
                api_key=openrouter_api_key,
                http_client=httpx.Client(),
            )
            model = engine.cfg.llm.model
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=100,
                temperature=0.2,
                timeout=5.0,
                extra_headers={
                    "HTTP-Referer": "https://jspark.internal",
                    "X-Title": "JSpark RMG Filler",
                },
            )
            reasoning = response.choices[0].message.content.strip()
            reasoning = reasoning.replace('"', '').replace('**', '')
            if reasoning:
                return reasoning
        except Exception as e:
            print(f"        OpenRouter failed: {e}. Falling back to Ollama...")

    # Second attempt / Fallback: Ollama
    if getattr(engine, "use_ollama", False):
        try:
            print("        Trying Ollama...")
            client = OpenAI(
                base_url=engine.cfg.llm.api_base,
                api_key="ollama",
                http_client=httpx.Client(),
            )
            response = client.chat.completions.create(
                model=engine.cfg.llm.model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=100,
                temperature=0.2,
                timeout=5.0,
            )
            reasoning = response.choices[0].message.content.strip()
            reasoning = reasoning.replace('"', '').replace('**', '')
            if reasoning:
                return reasoning
        except Exception as e:
            print(f"        Ollama failed: {e}. Falling back to rule-based string...")

    # Hardcoded Fallback
    if is_hire:
        return f"No internal candidates available at Tier {req_tier} matching requirements. External sourcing required."
    elif is_extend:
        return f"Project start date extension recommended to wait for internal resources."
    else:
        if not has_skills:
            return f"Allocated simply because they are available and no skillset constraints were given."
        else:
            sim = opt.score_breakdown.get("semantic_similarity", 0.0) if opt else 0.0
            return f"Selected {opt.recommended_employee_id if opt else ''} based on optimal semantic score of {sim:.3f}."


def sort_forecast_sheet(ws):
    """
    Group Forecast worksheet rows into project blocks and sort them.
    Primary key: Priority rank globally.
    Secondary key: Client Name alphabetically.
    """
    blocks = []
    current_block = []
    
    # Read values and copy all styles to memory first
    for r_idx in range(2, ws.max_row + 1):
        row_data = []
        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell_info = {
                "value": cell.value,
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "border": copy(cell.border) if cell.border else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format,
                "protection": copy(cell.protection) if cell.protection else None,
                "has_style": cell.has_style
            }
            row_data.append(cell_info)
            
        client_val = row_data[5]["value"] # Column F
        
        if client_val and str(client_val).strip():
            if current_block:
                blocks.append(current_block)
            current_block = []
            
        current_block.append((row_data, ws.row_dimensions[r_idx].height))
        
    if current_block:
        blocks.append(current_block)
        
    # Sort blocks globally by priority score, then client name
    def block_key(b):
        first_row_data = b[0][0]
        client_name = str(first_row_data[5]["value"] or "").strip().lower()
        
        sow_val = str(first_row_data[20]["value"] or "").strip().lower()
        sow_signed = sow_val in ("yes", "true", "1", "y")
        
        client_prio = str(first_row_data[4]["value"] or "").strip().lower()
        req_prio = str(first_row_data[12]["value"] or "").strip().lower()
        
        sow_group = 0 if sow_signed else 1
        
        if client_prio == "gold":
            prio_score = 1
        elif req_prio == "urgent":
            prio_score = 2
        elif client_prio == "silver":
            prio_score = 3
        elif req_prio == "high":
            prio_score = 4
        elif client_prio == "bronze":
            prio_score = 5
        elif req_prio == "medium":
            prio_score = 6
        elif client_prio == "other" or client_prio == "others":
            prio_score = 7
        elif req_prio == "low":
            prio_score = 8
        else:
            prio_score = 9
            
        return (sow_group, prio_score, client_name)
        
    blocks.sort(key=block_key)
    
    # Unmerge all merged cell ranges first to prevent delete_rows corruption
    merged_ranges = list(ws.merged_cells.ranges)
    for rng in merged_ranges:
        ws.unmerge_cells(rng.coord)
        
    # Delete existing rows
    ws.delete_rows(2, ws.max_row + 10)
    
    # Write back from memory using precise row indices and re-merge blocks
    new_r_idx = 2
    for block in blocks:
        start_row = new_r_idx
        end_row = new_r_idx + len(block) - 1
        
        for row_data, row_height in block:
            if row_height is not None:
                ws.row_dimensions[new_r_idx].height = row_height
            for c_idx, cell_info in enumerate(row_data, start=1):
                new_cell = ws.cell(row=new_r_idx, column=c_idx)
                new_cell.value = cell_info["value"]
                if cell_info["has_style"]:
                    if cell_info["font"]: new_cell.font = cell_info["font"]
                    if cell_info["fill"]: new_cell.fill = cell_info["fill"]
                    if cell_info["border"]: new_cell.border = cell_info["border"]
                    if cell_info["alignment"]: new_cell.alignment = cell_info["alignment"]
                    new_cell.number_format = cell_info["number_format"]
                    if cell_info["protection"]: new_cell.protection = cell_info["protection"]
            new_r_idx += 1
            
        # Re-merge the block rows for specific columns G, I, D, K, F, J, B, E, L
        if start_row < end_row:
            for col_letter in ['B', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L']:
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")


def run_excel_generation(engine, client_filter=None, project_filter=None):
    # Load reasoning flag from environment variable (default: True)
    do_reasoning = os.getenv("REASONING", "True").strip().lower() in ("true", "1", "yes")

    print("=" * 60)
    print("CoLab Resource Allocations Excel Filler")
    print("=" * 60)

    # 1b. Check reachability of LLM APIs to avoid slow connection hangs
    print("[1b/4] Checking LLM API reachability...")
    use_openrouter = False
    use_ollama = False

    openrouter_api_key = os.getenv("OPENROUTER_API_KEY", "").strip()
    if openrouter_api_key:
        try:
            import httpx
            print("      Testing OpenRouter connectivity...")
            resp = httpx.get("https://openrouter.ai/api/v1/models", headers={"Authorization": f"Bearer {openrouter_api_key}"}, timeout=2.0)
            if resp.status_code == 200:
                use_openrouter = True
                print("      OpenRouter API is reachable and API key is valid.")
            else:
                print(f"      OpenRouter API returned status {resp.status_code}. Disabling OpenRouter.")
        except Exception as e:
            print(f"      OpenRouter API is unreachable ({e}). Disabling OpenRouter.")

    try:
        import httpx
        print("      Testing local Ollama connectivity...")
        resp = httpx.get(engine.cfg.llm.api_base.rstrip("/") + "/tags", timeout=1.0)
        if resp.status_code in (200, 404):
            use_ollama = True
            print("      Ollama local API is reachable.")
    except Exception as e:
        print(f"      Ollama local API is unreachable. Disabling Ollama.")

    engine.use_openrouter = use_openrouter
    engine.use_ollama = use_ollama

    # 2. Run recommendations
    if client_filter:
        print(f"[2/4] Running recommendations for client: '{client_filter}'...")
    elif project_filter:
        print(f"[2/4] Running recommendations for project: '{project_filter}'...")
    else:
        print("[2/4] Running recommendations for all pipeline projects...")
    results = engine.run_all(client_filter=client_filter, project_filter=project_filter)
    print(f"      Calculated plans for {len(results)} projects.")

    # 3. Map (pipeline_id, project_role_index) to recommendations
    allocations_raw = []
    for res in results:
        plan = res.project_plan
        
        # Get matching pipeline project row to extract metadata for LLM
        matches = engine.ds.pipeline_projects[
            engine.ds.pipeline_projects["pipeline_id"] == plan.pipeline_id
        ]
        project_row = matches.iloc[0] if not matches.empty else {}
 
        # Get the pipeline roles for this project
        project_roles = engine.ds.pipeline_roles[
            engine.ds.pipeline_roles["pipeline_id"] == plan.pipeline_id
        ]

        # Within this project, match roles by their sequential position
        for r_idx, rp in enumerate(plan.role_plans):
            opt = rp.recommended_option
            if opt:
                emp_id = opt.recommended_employee_id
                cap = opt.available_capacity_pct
                
                # Format hires or wait lists or project extensions
                if opt.plan_type == 'D_HIRE':
                    emp_id = "HIRE"
                elif opt.plan_type == 'E_EXTEND_START':
                    emp_id = "EXTEND START DATE"
                elif opt.plan_type == 'C_WAIT':
                    emp_id = f"{emp_id or 'WAIT'} (Wait)"
                
                if not emp_id or str(emp_id).strip() == "" or str(emp_id).upper() == "NONE":
                    if opt.plan_type == 'D_HIRE':
                        emp_id = "HIRE"
                    else:
                        emp_id = "EXTEND START DATE"
            else:
                emp_id = "EXTEND START DATE"
                cap = 0.0
                
            # Calculate skillset match status
            role_row = project_roles.iloc[r_idx] if r_idx < len(project_roles) else {}
            req_skills = role_row.get("skillset_notes", "") if not isinstance(role_row, dict) else ""
            has_req_skills = bool(req_skills and len(str(req_skills).strip()) > 2)
            
            skillset_match = "No"
            if has_req_skills:
                if opt and opt.plan_type != 'D_HIRE' and opt.recommended_employee_id:
                    sim = opt.score_breakdown.get("semantic_similarity", 0.0) if opt.score_breakdown else 0.0
                    if sim >= 0.50:
                        skillset_match = "Complete"
                    elif sim > 0.0:
                        skillset_match = "Partial"
                    else:
                        skillset_match = "No"
            
            allocations_raw.append({
                "pipeline_id": plan.pipeline_id,
                "role_name": rp.role_name,
                "r_idx": r_idx,
                "emp_id": emp_id,
                "cap": cap,
                "skillset_match": skillset_match,
                "project_row": project_row,
                "rp": rp,
                "opt": opt,
                "req_skills": req_skills
            })

    # Now, run reasoning generation in parallel!
    allocations = {}
    if do_reasoning:
        print(f"      Generating reasoning for {len(allocations_raw)} roles in parallel...")
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        def run_one(item):
            res_str = generate_reasoning(engine, item["project_row"], item["rp"], item["opt"], item["req_skills"])
            return item["pipeline_id"], item["r_idx"], item["emp_id"], item["cap"], res_str, item["skillset_match"]

        # Use 15 threads for high concurrency
        with ThreadPoolExecutor(max_workers=15) as executor:
            futures = [executor.submit(run_one, item) for item in allocations_raw]
            
            count = 0
            for fut in as_completed(futures):
                pid, r_idx, emp_id, cap, reasoning, skillset_match = fut.result()
                allocations[(pid, r_idx)] = (emp_id, cap, reasoning, skillset_match)
                count += 1
                if count % 20 == 0 or count == len(allocations_raw):
                    print(f"        Progress: {count}/{len(allocations_raw)} reasoning notes generated.")
    else:
        print("      Mapping recommendations (skipping reasoning)...")
        for item in allocations_raw:
            allocations[(item["pipeline_id"], item["r_idx"])] = (item["emp_id"], item["cap"], "", item["skillset_match"])

    print(f"      Mapped {len(allocations)} role recommendations.")

    # 4. Read input Excel and fill columns
    input_path = backend_dir / "data" / "07__260624_Pipeline_Details.xlsx"
    print(f"[3/4] Reading pipeline details workbook from {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["Forecast"]

    # Perform blocks sorting on Forecast worksheet
    # print("      Sorting Forecast project blocks...")
    # sort_forecast_sheet(ws)

    print("      Processing rows and filling recommended resources...")
    current_pipeline_id = None
    project_role_counter = 0
    current_project = False

    for r_idx in range(2, ws.max_row + 1):
        client_val = ws.cell(row=r_idx, column=6).value
        resource_val = ws.cell(row=r_idx, column=15).value

        # Clean formula objects if any
        if hasattr(client_val, "value"):
            client_val = None
        if hasattr(resource_val, "value"):
            resource_val = None

        if client_val and str(client_val).strip():
            current_project = True
            
            # Reconstruct pipeline_id exactly as done in generate_datacube.py
            cluster = ws.cell(row=r_idx, column=1).value
            likely_start_raw = ws.cell(row=r_idx, column=8).value
            
            if hasattr(cluster, "value"):
                cluster = None
            if hasattr(likely_start_raw, "value"):
                likely_start_raw = None
                
            cluster_str = str(int(cluster)) if cluster is not None and str(cluster).replace('.0','').isdigit() else "0"
            
            likely_start_str = "TBD"
            if likely_start_raw:
                try:
                    likely_start_dt = pd.to_datetime(str(likely_start_raw).split(" ")[0], dayfirst=True, errors="coerce")
                    if not pd.isna(likely_start_dt):
                        likely_start_str = likely_start_dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
            
            client_name = str(client_val).strip().replace(" ", "_")
            current_pipeline_id = f"{client_name}_{cluster_str}_{likely_start_str}"
            project_role_counter = 0

        if not current_project:
            continue

        if not resource_val or not str(resource_val).strip():
            continue

        # This row is a valid role slot inside current_pipeline_id
        key = (current_pipeline_id, project_role_counter)
        if key in allocations:
            emp_id, cap, reasoning, skillset_match = allocations[key]
            ws.cell(row=r_idx, column=17, value=emp_id)  # Resource Recommended (Col Q)
            ws.cell(row=r_idx, column=18, value=cap)     # % Available (Col R)
            ws.cell(row=r_idx, column=20, value=skillset_match) # Skillset Match (Col T)
            ws.cell(row=r_idx, column=22, value=reasoning) # Comments (Col V)

        project_role_counter += 1

    # 5. Save the updated workbook
    output_dir = backend_dir / "output_excels"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "updated_resources.xlsx"
    
    print(f"[4/4] Saving filled workbook to {output_path}...")
    try:
        wb.save(output_path)
        wb.close()
        print("\nSuccess! Excel file filled successfully.")
    except PermissionError:
        fallback_path = output_dir / "updated_resources_temp.xlsx"
        print(f"\nWarning: Permission Denied while saving to {output_path}.")
        print(f"Trying to save to fallback location: {fallback_path}...")
        try:
            wb.save(fallback_path)
            wb.close()
            print(f"\nSuccess! Excel file filled successfully at: {fallback_path}")
        except PermissionError:
            print(f"\nERROR: Permission Denied while saving to {fallback_path} as well.")
            raise PermissionError("Could not save to template Excel path due to file lock.")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="CoLab Resource Allocations Excel Filler")
    parser.add_argument("--client", type=str, help="Filter by client name (case-insensitive, e.g. --client KIA)")
    parser.add_argument("--project", type=str, help="Filter by specific pipeline project ID (case-insensitive, e.g. --project KIA_5_2026-06-22)")
    args = parser.parse_args()

    # 1. Initialize recommendation engine
    print("[1/4] Initializing Recommendation Engine...")
    engine = RecommendationEngine(DEFAULT_CONFIG)

    run_excel_generation(engine, client_filter=args.client, project_filter=args.project)

if __name__ == "__main__":
    main()
